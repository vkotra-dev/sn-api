from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
import re
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_HEADERS = {"plot no", "dim ft", "size ft"}


@dataclass(slots=True)
class PlotMetadata:
    plot_no: str
    owner: str | None
    dim_ft: str | None
    dim_type: str | None
    area_sq_ft: float | None
    area_sq_yd: float | None
    facing: str | None


def _normalize_header(value) -> str:
    return str(value).strip().lower()


def _normalize_plot_no(value) -> str:
    if value is None:
        raise ValueError("Plot no is missing")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    if not text:
        raise ValueError("Plot no is missing")
    try:
        number = float(text)
    except ValueError:
        return text
    if number.is_integer():
        return str(int(number))
    return text


def parse_dimension(dim_str: str | None) -> str | None:
    if dim_str is None:
        return None
    value = str(dim_str).strip()
    if not value:
        return None
    if "," in value:
        match = re.match(r"([\d.]+),([\d.]+)\*([\d.]+)", value)
        if match:
            return "trap"
    else:
        match = re.match(r"([\d.]+)\*([\d.]+)", value)
        if match:
            return "rect"
    return None


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_excel_metadata(excel_path: Path) -> dict[str, PlotMetadata]:
    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook.active

    headers = [_normalize_header(cell.value) for cell in sheet[1]]
    header_index = {header: index for index, header in enumerate(headers)}
    missing = REQUIRED_HEADERS - set(header_index)
    if missing:
        raise ValueError(f"Excel file missing required columns: {', '.join(sorted(missing))}")

    metadata: dict[str, PlotMetadata] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None and cell != "" for cell in row):
            continue

        plot_no = _normalize_plot_no(row[header_index["plot no"]])
        dim_ft = row[header_index["dim ft"]]
        size_ft = row[header_index["size ft"]]
        owner = row[header_index["owner"]] if "owner" in header_index else None
        facing = row[header_index["facing"]] if "facing" in header_index else None
        size_yd = row[header_index["size yards"]] if "size yards" in header_index else None

        dim_ft_text = None if dim_ft is None else str(dim_ft).strip() or None
        if dim_ft_text is None:
            raise ValueError(f"Plot {plot_no} is missing Dim ft")

        area_sq_ft = _to_float(size_ft)
        if area_sq_ft is None:
            raise ValueError(f"Plot {plot_no} is missing Size ft")

        area_sq_yd = _to_float(size_yd)
        if area_sq_yd is None:
            area_sq_yd = round(area_sq_ft / 9, 2)

        if plot_no in metadata:
            raise ValueError(f"Duplicate plot number in Excel: {plot_no}")

        metadata[plot_no] = PlotMetadata(
            plot_no=plot_no,
            owner=None if owner is None else str(owner).strip() or None,
            dim_ft=dim_ft_text,
            dim_type=parse_dimension(dim_ft_text),
            area_sq_ft=area_sq_ft,
            area_sq_yd=area_sq_yd,
            facing=None if facing is None else str(facing).strip() or None,
        )

    if not metadata:
        raise ValueError("Excel file does not contain any plot rows")

    return metadata

